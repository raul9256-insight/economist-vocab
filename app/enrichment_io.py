from __future__ import annotations

import csv
import io
import json
import sqlite3
from pathlib import Path
from typing import BinaryIO

import openpyxl

from app.db import definitions_for_word, parts_of_speech_for_word
from economist_vocab import normalize_word


EXPECTED_COLUMNS = [
    "lemma",
    "band_label",
    "parts_of_speech",
    "chinese_definitions",
    "pronunciation",
    "english_definition",
    "example_sentence",
    "synonyms",
    "sentence_distractors",
    "notes",
]

AI_POWER_EXPECTED_COLUMNS = [
    "completion_status",
    "missing_fields",
    "category_slug",
    "category_name",
    "english",
    "type_of_word",
    "english_definition",
    "traditional_chinese",
    "simplified_chinese",
    "example_sentence",
    "ai_prompt_example",
    "prompt_strategic",
    "prompt_creative",
    "prompt_technical",
    "prompt_finance",
    "prompt_education",
    "ipa",
    "notes",
]

TAXONOMY_EXPECTED_COLUMNS = [
    "cluster_slug",
    "cluster_label",
    "core_meaning",
    "cluster_domain",
    "lemma",
    "band_label",
    "role",
    "stage_rank",
    "stage_label",
    "cluster_note",
    "formality_level",
    "precision_level",
    "exam_relevance",
    "business_relevance",
    "ai_relevance",
    "productivity_likelihood",
    "word_domain",
    "register_note",
    "usage_note",
    "related_word",
    "relation_type",
    "relation_explanation",
    "relation_strength",
]


def parse_list_field(value: str) -> list[str]:
    if not value:
        return []
    chunks = str(value).replace(";", "\n").splitlines()
    return [chunk.strip() for chunk in chunks if chunk.strip()]


def iter_import_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [{(key or "").strip(): (value or "").strip() for key, value in row.items()} for row in reader]
    if suffix in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        result: list[dict[str, str]] = []
        for raw in rows[1:]:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = raw[index] if index < len(raw) else ""
                row[header] = "" if value is None else str(value).strip()
            result.append(row)
        return result
    raise ValueError("Only .csv and .xlsx files are supported.")


def export_template(
    conn: sqlite3.Connection,
    output_path: Path,
    *,
    band_rank: int | None = None,
    limit: int | None = None,
    missing_only: bool = False,
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    clauses = ["1 = 1"]
    params: list[object] = []
    if band_rank is not None:
        clauses.append("words.best_band_rank = ?")
        params.append(band_rank)
    if missing_only:
        clauses.append("(word_enrichment.english_definition IS NULL OR word_enrichment.english_definition = '' OR word_enrichment.example_sentence IS NULL OR word_enrichment.example_sentence = '')")
    limit_clause = f"LIMIT {int(limit)}" if limit else ""
    rows = conn.execute(
        f"""
        SELECT words.id, words.lemma, words.best_band_label,
               COALESCE(word_enrichment.pronunciation, '') AS pronunciation,
               COALESCE(word_enrichment.english_definition, '') AS english_definition,
               COALESCE(word_enrichment.example_sentence, '') AS example_sentence,
               COALESCE(word_enrichment.synonyms_json, '[]') AS synonyms_json,
               COALESCE(word_enrichment.sentence_distractors_json, '[]') AS sentence_distractors_json,
               COALESCE(study_cards.notes, '') AS notes
        FROM words
        LEFT JOIN word_enrichment ON word_enrichment.word_id = words.id
        LEFT JOIN study_cards ON study_cards.word_id = words.id
        WHERE {' AND '.join(clauses)}
        ORDER BY words.best_band_rank, words.lemma
        {limit_clause}
        """,
        params,
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "enrichment"
    ws.append(EXPECTED_COLUMNS)
    for row in rows:
        ws.append(
            [
                row["lemma"],
                row["best_band_label"],
                ", ".join(parts_of_speech_for_word(conn, row["id"])),
                " | ".join(definitions_for_word(conn, row["id"])),
                row["pronunciation"],
                row["english_definition"],
                row["example_sentence"],
                "\n".join(json.loads(row["synonyms_json"])),
                "\n".join(json.loads(row["sentence_distractors_json"])),
                row["notes"],
            ]
        )
    wb.save(output_path)
    return len(rows)


AI_POWER_COMPLETION_FIELDS = [
    "type_of_word",
    "english_definition",
    "traditional_chinese",
    "simplified_chinese",
    "example_sentence",
    "ai_prompt_example",
    "ipa",
]


def export_ai_power_template(categories: list[dict], output_path: Path, *, missing_only: bool = False) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ai_power_vocab"
    ws.append(AI_POWER_EXPECTED_COLUMNS)

    total_rows = 0
    for category in categories:
        entries = category.get("entries")
        if entries is None:
            entries = [{"english": term} for term in category.get("terms", [])]
        for entry in sorted(entries, key=lambda value: value.get("english", "").lower()):
            missing_fields = [
                field
                for field in AI_POWER_COMPLETION_FIELDS
                if not str(entry.get(field, "")).strip()
            ]
            if missing_only and not missing_fields:
                continue
            completion_status = "complete" if not missing_fields else "missing"
            ws.append(
                [
                    completion_status,
                    ", ".join(missing_fields),
                    category.get("slug", ""),
                    category.get("english_title", category.get("title", "")),
                    entry.get("english", ""),
                    entry.get("type_of_word", ""),
                    entry.get("english_definition", ""),
                    entry.get("traditional_chinese", ""),
                    entry.get("simplified_chinese", ""),
                    entry.get("example_sentence", "") or category.get("normal_example", ""),
                    entry.get("ai_prompt_example", "") or category.get("prompt_example", ""),
                    entry.get("prompt_strategic", ""),
                    entry.get("prompt_creative", ""),
                    entry.get("prompt_technical", ""),
                    entry.get("prompt_finance", ""),
                    entry.get("prompt_education", ""),
                    entry.get("ipa", ""),
                    entry.get("notes", ""),
                ]
            )
            total_rows += 1
        new_term_slots = max(0, int(category.get("target_count", 0) or 0) - len(entries))
        for _ in range(new_term_slots):
            ws.append(
                [
                    "new_term_needed",
                    "english, " + ", ".join(AI_POWER_COMPLETION_FIELDS),
                    category.get("slug", ""),
                    category.get("english_title", category.get("title", "")),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Add one new AI Power vocabulary item for this category.",
                ]
            )
            total_rows += 1

    wb.save(output_path)
    return total_rows


def export_taxonomy_template(
    conn: sqlite3.Connection,
    output_path: Path,
    *,
    band_rank: int | None = None,
    limit: int | None = None,
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    clauses = ["1 = 1"]
    params: list[object] = []
    if band_rank is not None:
        clauses.append("words.best_band_rank = ?")
        params.append(band_rank)
    limit_clause = f"LIMIT {int(limit)}" if limit else ""
    rows = conn.execute(
        f"""
        SELECT
            words.id,
            words.lemma,
            words.best_band_label,
            COALESCE(vocab_clusters.slug, '') AS cluster_slug,
            COALESCE(vocab_clusters.label, '') AS cluster_label,
            COALESCE(vocab_clusters.core_meaning, '') AS core_meaning,
            COALESCE(vocab_clusters.domain, '') AS cluster_domain,
            COALESCE(vocab_cluster_words.role, '') AS role,
            COALESCE(vocab_cluster_words.stage_rank, 1) AS stage_rank,
            COALESCE(vocab_cluster_words.stage_label, '') AS stage_label,
            COALESCE(vocab_cluster_words.note, '') AS cluster_note,
            COALESCE(word_progression_attributes.formality_level, 1) AS formality_level,
            COALESCE(word_progression_attributes.precision_level, 1) AS precision_level,
            COALESCE(word_progression_attributes.exam_relevance, 0) AS exam_relevance,
            COALESCE(word_progression_attributes.business_relevance, 0) AS business_relevance,
            COALESCE(word_progression_attributes.ai_relevance, 0) AS ai_relevance,
            COALESCE(word_progression_attributes.productivity_likelihood, 0) AS productivity_likelihood,
            COALESCE(word_progression_attributes.domain, '') AS word_domain,
            COALESCE(word_progression_attributes.register_note, '') AS register_note,
            COALESCE(word_progression_attributes.usage_note, '') AS usage_note
        FROM words
        LEFT JOIN vocab_cluster_words ON vocab_cluster_words.word_id = words.id
        LEFT JOIN vocab_clusters ON vocab_clusters.id = vocab_cluster_words.cluster_id
        LEFT JOIN word_progression_attributes ON word_progression_attributes.word_id = words.id
        WHERE {' AND '.join(clauses)}
        ORDER BY words.best_band_rank DESC, words.lemma
        {limit_clause}
        """,
        params,
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "taxonomy"
    ws.append(TAXONOMY_EXPECTED_COLUMNS)
    for row in rows:
        ws.append(
            [
                row["cluster_slug"],
                row["cluster_label"],
                row["core_meaning"],
                row["cluster_domain"],
                row["lemma"],
                row["best_band_label"],
                row["role"],
                row["stage_rank"],
                row["stage_label"],
                row["cluster_note"],
                row["formality_level"],
                row["precision_level"],
                row["exam_relevance"],
                row["business_relevance"],
                row["ai_relevance"],
                row["productivity_likelihood"],
                row["word_domain"],
                row["register_note"],
                row["usage_note"],
                "",
                "",
                "",
                "",
            ]
        )
    wb.save(output_path)
    return len(rows)


def import_ai_power_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], dict[str, int]]:
    stats = {"updated": 0, "skipped": 0}
    result: list[dict[str, str]] = []
    for raw_row in rows:
        english = (
            raw_row.get("english", "")
            or raw_row.get("Vocabulary", "")
            or raw_row.get("vocabulary", "")
            or raw_row.get("lemma", "")
        ).strip()
        if not english:
            stats["skipped"] += 1
            continue
        result.append(
            {
                "category_slug": (raw_row.get("category_slug", "") or raw_row.get("Category Slug", "")).strip(),
                "category_name": (raw_row.get("category_name", "") or raw_row.get("Category Name", "")).strip(),
                "english": english,
                "type_of_word": (
                    raw_row.get("type_of_word", "")
                    or raw_row.get("Type of word", "")
                    or raw_row.get("part_of_speech", "")
                ).strip(),
                "english_definition": (
                    raw_row.get("english_definition", "")
                    or raw_row.get("English Definition", "")
                ).strip(),
                "traditional_chinese": (
                    raw_row.get("traditional_chinese", "")
                    or raw_row.get("Traditional Chinese", "")
                ).strip(),
                "simplified_chinese": (
                    raw_row.get("simplified_chinese", "")
                    or raw_row.get("Simplified Chinese", "")
                ).strip(),
                "example_sentence": (
                    raw_row.get("example_sentence", "")
                    or raw_row.get("Example Sentence", "")
                ).strip(),
                "ai_prompt_example": (
                    raw_row.get("ai_prompt_example", "")
                    or raw_row.get("AI Prompt Example", "")
                ).strip(),
                "prompt_strategic": (
                    raw_row.get("prompt_strategic", "")
                    or raw_row.get("Strategic Prompt", "")
                    or raw_row.get("專業諮詢與管理", "")
                ).strip(),
                "prompt_creative": (
                    raw_row.get("prompt_creative", "")
                    or raw_row.get("Creative Prompt", "")
                    or raw_row.get("創意與內容行銷", "")
                ).strip(),
                "prompt_technical": (
                    raw_row.get("prompt_technical", "")
                    or raw_row.get("Technical Prompt", "")
                    or raw_row.get("技術、工程與學術", "")
                ).strip(),
                "prompt_finance": (
                    raw_row.get("prompt_finance", "")
                    or raw_row.get("Finance Prompt", "")
                    or raw_row.get("金融、法律與合規", "")
                ).strip(),
                "prompt_education": (
                    raw_row.get("prompt_education", "")
                    or raw_row.get("Education Prompt", "")
                    or raw_row.get("教育與終身學習", "")
                ).strip(),
                "ipa": (raw_row.get("ipa", "") or raw_row.get("IPA", "") or raw_row.get("pronunciation", "")).strip(),
                "notes": (raw_row.get("notes", "") or raw_row.get("Notes", "")).strip(),
            }
        )
        stats["updated"] += 1
    return result, stats


def import_enrichment_rows(conn: sqlite3.Connection, rows: list[dict[str, str]]) -> dict[str, int]:
    stats = {"updated": 0, "skipped": 0, "missing_words": 0}
    for raw_row in rows:
        lemma = normalize_word(raw_row.get("lemma", ""))
        if not lemma:
            stats["skipped"] += 1
            continue
        word = conn.execute(
            "SELECT id FROM words WHERE normalized_lemma = ?",
            (lemma,),
        ).fetchone()
        if word is None:
            stats["missing_words"] += 1
            continue
        current = conn.execute(
            """
            SELECT pronunciation, english_definition, synonyms_json, example_sentence, sentence_distractors_json
            FROM word_enrichment
            WHERE word_id = ?
            """,
            (word["id"],),
        ).fetchone()
        current_pronunciation = current["pronunciation"] if current else ""
        current_english = current["english_definition"] if current else ""
        current_synonyms = json.loads(current["synonyms_json"]) if current else []
        current_example = current["example_sentence"] if current else ""
        current_distractors = json.loads(current["sentence_distractors_json"]) if current else []

        incoming_pronunciation = (raw_row.get("pronunciation", "") or raw_row.get("ipa", "")).strip()
        incoming_english = raw_row.get("english_definition", "").strip()
        incoming_example = raw_row.get("example_sentence", "").strip()
        incoming_synonyms = parse_list_field(raw_row.get("synonyms", ""))
        incoming_distractors = parse_list_field(raw_row.get("sentence_distractors", ""))
        notes = raw_row.get("notes", "").strip()
        has_new_content = any([incoming_pronunciation, incoming_english, incoming_example, incoming_synonyms, incoming_distractors, notes])
        if not has_new_content:
            stats["skipped"] += 1
            continue

        pronunciation = incoming_pronunciation or current_pronunciation
        english_definition = incoming_english or current_english
        example_sentence = incoming_example or current_example
        synonyms = incoming_synonyms or current_synonyms
        sentence_distractors = incoming_distractors or current_distractors

        conn.execute(
            """
            INSERT INTO word_enrichment (
                word_id, pronunciation, english_definition, synonyms_json, example_sentence, sentence_distractors_json
            )
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(word_id) DO UPDATE SET
                pronunciation = excluded.pronunciation,
                english_definition = excluded.english_definition,
                synonyms_json = excluded.synonyms_json,
                example_sentence = excluded.example_sentence,
                sentence_distractors_json = excluded.sentence_distractors_json,
                updated_at = CURRENT_TIMESTAMP
            """,
            (
                word["id"],
                pronunciation,
                english_definition,
                json.dumps(synonyms, ensure_ascii=False),
                example_sentence,
                json.dumps(sentence_distractors, ensure_ascii=False),
            ),
        )
        if notes:
            conn.execute(
                """
                UPDATE study_cards
                SET notes = ?, updated_at = CURRENT_TIMESTAMP
                WHERE word_id = ?
                """,
                (notes, word["id"]),
            )
        stats["updated"] += 1
    conn.commit()
    return stats


def import_taxonomy_rows(conn: sqlite3.Connection, rows: list[dict[str, str]]) -> dict[str, int]:
    stats = {"updated": 0, "skipped": 0, "missing_words": 0, "missing_related_words": 0}
    cluster_cache: dict[str, int] = {}

    def as_int(value: str, default: int) -> int:
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return default

    for raw_row in rows:
        lemma = normalize_word(raw_row.get("lemma", "") or raw_row.get("Vocabulary", "") or raw_row.get("vocabulary", ""))
        if not lemma:
            stats["skipped"] += 1
            continue
        word = conn.execute(
            "SELECT id FROM words WHERE normalized_lemma = ?",
            (lemma,),
        ).fetchone()
        if word is None:
            stats["missing_words"] += 1
            continue

        cluster_slug = (raw_row.get("cluster_slug", "") or raw_row.get("Cluster Slug", "")).strip()
        cluster_label = (raw_row.get("cluster_label", "") or raw_row.get("Cluster Label", "")).strip()
        core_meaning = (raw_row.get("core_meaning", "") or raw_row.get("Core Meaning", "")).strip()
        cluster_domain = (raw_row.get("cluster_domain", "") or raw_row.get("Cluster Domain", "")).strip()
        role = (raw_row.get("role", "") or raw_row.get("Role", "")).strip() or "member"
        stage_rank = as_int(raw_row.get("stage_rank", "") or raw_row.get("Stage Rank", ""), 1)
        stage_label = (raw_row.get("stage_label", "") or raw_row.get("Stage Label", "")).strip()
        cluster_note = (raw_row.get("cluster_note", "") or raw_row.get("Cluster Note", "")).strip()

        if cluster_slug and cluster_label:
            cluster_id = cluster_cache.get(cluster_slug)
            if cluster_id is None:
                conn.execute(
                    """
                    INSERT INTO vocab_clusters (slug, label, core_meaning, domain)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(slug) DO UPDATE SET
                        label = excluded.label,
                        core_meaning = excluded.core_meaning,
                        domain = excluded.domain
                    """,
                    (cluster_slug, cluster_label, core_meaning, cluster_domain),
                )
                cluster = conn.execute("SELECT id FROM vocab_clusters WHERE slug = ?", (cluster_slug,)).fetchone()
                cluster_id = cluster["id"]
                cluster_cache[cluster_slug] = cluster_id
            conn.execute(
                """
                INSERT INTO vocab_cluster_words (cluster_id, word_id, role, stage_rank, stage_label, note)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(cluster_id, word_id) DO UPDATE SET
                    role = excluded.role,
                    stage_rank = excluded.stage_rank,
                    stage_label = excluded.stage_label,
                    note = excluded.note
                """,
                (cluster_id, word["id"], role, stage_rank, stage_label, cluster_note),
            )

        conn.execute(
            """
            INSERT INTO word_progression_attributes (
                word_id, formality_level, precision_level, exam_relevance,
                business_relevance, ai_relevance, productivity_likelihood,
                domain, register_note, usage_note
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(word_id) DO UPDATE SET
                formality_level = excluded.formality_level,
                precision_level = excluded.precision_level,
                exam_relevance = excluded.exam_relevance,
                business_relevance = excluded.business_relevance,
                ai_relevance = excluded.ai_relevance,
                productivity_likelihood = excluded.productivity_likelihood,
                domain = excluded.domain,
                register_note = excluded.register_note,
                usage_note = excluded.usage_note,
                updated_at = CURRENT_TIMESTAMP
            """,
            (
                word["id"],
                as_int(raw_row.get("formality_level", "") or raw_row.get("Formality Level", ""), 1),
                as_int(raw_row.get("precision_level", "") or raw_row.get("Precision Level", ""), 1),
                as_int(raw_row.get("exam_relevance", "") or raw_row.get("Exam Relevance", ""), 0),
                as_int(raw_row.get("business_relevance", "") or raw_row.get("Business Relevance", ""), 0),
                as_int(raw_row.get("ai_relevance", "") or raw_row.get("AI Relevance", ""), 0),
                as_int(raw_row.get("productivity_likelihood", "") or raw_row.get("Productivity Likelihood", ""), 0),
                (raw_row.get("word_domain", "") or raw_row.get("Word Domain", "")).strip(),
                (raw_row.get("register_note", "") or raw_row.get("Register Note", "")).strip(),
                (raw_row.get("usage_note", "") or raw_row.get("Usage Note", "")).strip(),
            ),
        )

        related_word = normalize_word(raw_row.get("related_word", "") or raw_row.get("Related Word", ""))
        relation_type = (raw_row.get("relation_type", "") or raw_row.get("Relation Type", "")).strip()
        if related_word and relation_type:
            target = conn.execute(
                "SELECT id FROM words WHERE normalized_lemma = ?",
                (related_word,),
            ).fetchone()
            if target is None:
                stats["missing_related_words"] += 1
            else:
                conn.execute(
                    """
                    INSERT INTO word_relationships (
                        source_word_id, target_word_id, relation_type, explanation, strength
                    )
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(source_word_id, target_word_id, relation_type) DO UPDATE SET
                        explanation = excluded.explanation,
                        strength = excluded.strength
                    """,
                    (
                        word["id"],
                        target["id"],
                        relation_type,
                        (raw_row.get("relation_explanation", "") or raw_row.get("Relation Explanation", "")).strip(),
                        as_int(raw_row.get("relation_strength", "") or raw_row.get("Relation Strength", ""), 1),
                    ),
                )

        stats["updated"] += 1

    conn.commit()
    return stats
